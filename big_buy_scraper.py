import time
import threading

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

import openpyxl

import tkinter as tk
from tkinter import ttk
from tkinter import messagebox


def read_links(from_row, to_row):
    wb = openpyxl.load_workbook('links.xlsx')
    sheet = wb['Sheet1']
    links = []
    if from_row == -1 and to_row == -1:
        from_row = 2
        to_row = sheet.max_row + 1

    for row in range(from_row, to_row + 1):
        link = sheet.cell(row=row, column=1).value
        if link is not None and link.strip() != '':
            links.append(link.strip())

    return links


def read_login():
    file = open('login.txt', 'r')
    email = file.readline().strip().replace('\n', '')
    password = file.readline().strip().replace('\n', '')
    file.close()
    return email, password


def login(email, password):
    WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.CSS_SELECTOR, ".small.login.item")))
    driver.find_element(By.CSS_SELECTOR, ".small.login.item").click()
    WebDriverWait(driver, 30).until(EC.visibility_of_element_located((By.CSS_SELECTOR, "#password")))
    time.sleep(1)
    driver.find_elements(By.CSS_SELECTOR, 'input[name="email"]')[1].send_keys(email)
    driver.find_element(By.CSS_SELECTOR, "#password").send_keys(password)
    time.sleep(1)
    driver.find_element(By.CSS_SELECTOR, 'button[type="submit"]').click()
    WebDriverWait(driver, 30).until(EC.invisibility_of_element_located((By.CSS_SELECTOR, "div[class='bb-popup-login']")))


def write_data(link, data, column):
    wb = openpyxl.load_workbook('big_buy_products.xlsx')

    sheet = wb['Sheet1']
    sheet.cell(row=1, column=column).value = link
    n_ref_nums = len(data[0])
    n_tags = len(data[1])

    for row in range(2, n_tags + 2):  # write tags
        sheet.cell(row=row, column=column).value = data[1][row - 2]

    for row in range(9, n_ref_nums + 9):  # write ref numbers
        sheet.cell(row=row, column=column).value = data[0][row - n_ref_nums - 9]

    wb.save('big_buy_products.xlsx')


def get_data_from_link(link):
    ref_nums = []
    tags = []

    driver.get(link)
    WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.CSS_SELECTOR, 'div[class="productList-item"]')))

    # scroll into view to the label
    driver.execute_script("arguments[0].scrollIntoView();", driver.find_element(By.CSS_SELECTOR, 'label[for="list_view"]'))
    time.sleep(3)
    driver.execute_script("window.scrollBy(0, -180);")
    time.sleep(3)
    driver.find_element(By.CSS_SELECTOR, 'label[for="list_view"]').click()  # change view to list
    time.sleep(2)

    more_products = True

    while more_products:

        products = driver.find_elements(By.CSS_SELECTOR, 'div[class="productList-item"]')
        for product in products:
            if product.find_element(By.CSS_SELECTOR, 'p[class="stockAvailability-title"]').text.strip().lower() == 'available':
                try:  # skip renewed products
                    if product.find_element(By.CSS_SELECTOR, 'div[class="productCard-ribbon"]').text.strip().lower() == 'renewed':
                        continue
                except:
                    pass

                ref = product.find_elements(By.CSS_SELECTOR, 'div[class="productCard-sizesItem"]')[0].text.strip().replace('Ref. ', '')
                ref_nums.append(ref)

                print(ref)
        try:  # check if there are more products
            more_products = driver.find_element(By.CSS_SELECTOR, 'button[data-js-paginador="next"]').is_enabled()
            driver.execute_script("arguments[0].scrollIntoView();", driver.find_element(By.CSS_SELECTOR, 'button[data-js-paginador="next"]'))
            time.sleep(3)
            driver.execute_script("window.scrollBy(0, -180);")
            time.sleep(3)
            driver.find_element(By.CSS_SELECTOR, 'button[data-js-paginador="next"]').click()
            time.sleep(2)
        except:
            more_products = False


    tags = driver.find_elements(By.CSS_SELECTOR, 'span[itemprop="title"]')
    tags = [tag.text.strip() for tag in tags]
    for i in range(len(tags)):
        tags[i] = ':' + tags[i]

    tags.pop(0)  # remove first element from tags list (home page)
    print(tags)

    return [ref_nums, tags]


def scrape_links(links):
    # create Excel file or clear it if it already exists
    wb = openpyxl.Workbook()
    # create sheet
    wb.create_sheet('Sheet1', 0)
    wb.save('big_buy_products.xlsx')

    i = 1
    for link in links:  # get ref number for each link and write it to the Excel file
        try:
            data = get_data_from_link(link)
            write_data(link, data, i)
        except Exception as e:
            print(e)

        i += 1


def main():
    global scraping, driver

    scraping = True
    status_label.config(text="Scraping...")
    status_label.configure(foreground="#2ecc71")

    try:
        options = webdriver.FirefoxOptions()
        if headless_var.get() == 1:
            options.add_argument("--headless")
        driver = webdriver.Firefox(executable_path='geckodriver.exe', firefox_binary=ff_path, options=options)
        driver.maximize_window()

    except Exception as e:
        messagebox.showerror("Error", "Can't open firefox driver.\n" + str(e))
        return

    try:
        from_row = from_row_entry.get().strip()
        to_row = to_row_entry.get().strip()
        if from_row == '' and to_row == '':
            from_row = -1
            to_row = -1
        else:
            from_row = int(from_row)
            to_row = int(to_row)
            assert 0 < from_row <= to_row and to_row > 0

    except Exception as e:
        messagebox.showerror("Error", "Invalid row numbers")
        return

    try:
        email, password = read_login()
        driver.get("https://www.bigbuy.eu/en/")
        login(email, password)
        links = read_links(from_row, to_row)
        scrape_links(links)
        messagebox.showinfo("Success", "Scraping completed successfully")

    except Exception as e:
        messagebox.showerror("Error", "Something went wrong.\n" + str(e))

    try:
        driver.quit()
    except:
        pass

    scraping = False
    status_label.config(text="Status: Idle")
    status_label.configure(foreground="#e74c3c")


def start_main_thread():
    if scraping:
        messagebox.showinfo("Info", "Scraping is already in progress")
        return

    main_thread = threading.Thread(target=main)
    main_thread.daemon = True
    main_thread.start()


# Create the main window
window = tk.Tk()
window.title("Big Buy Scraper")
window.configure(bg="#34495e")

# Define custom styles for labels and buttons
style = ttk.Style(window)
style.configure("TLabel", foreground="white", background="#34495e", font=("Arial", 12))
style.configure("TButton", foreground="#34495e", background="#2980b9", font=("Arial", 12))
style.map("TButton",
          foreground=[('pressed', '#34495e'), ('active', '#34495e')],
          background=[('pressed', '#3498db'), ('active', '#3498db')])

# Create and configure the input fields with rounded style
entry_style = ttk.Style()
entry_style.configure("Rounded.TEntry", fieldbackground="#ecf0f1", bordercolor="#bdc3c7", borderwidth=2,
                      relief="sunken", padding=5)
entry_style.map("Rounded.TEntry",
                fieldbackground=[('focus', '#ffffff'), ('!focus', '#ecf0f1')])

# Input fields in the first column
from_row_label = ttk.Label(window, text="From Row:")
from_row_label.grid(row=0, column=0, padx=10, pady=50, sticky="E")
from_row_entry = ttk.Entry(window, style="Rounded.TEntry")
from_row_entry.grid(row=0, column=1, padx=10, pady=50, sticky="W")

to_row_label = ttk.Label(window, text="To Row:")
to_row_label.grid(row=0, column=2, padx=10, pady=50, sticky="E")
to_row_entry = ttk.Entry(window, style="Rounded.TEntry")
to_row_entry.grid(row=0, column=3, padx=10, pady=50, sticky="W")

# headless browser checkbox
headless_var = tk.IntVar()
headless_checkbox = ttk.Checkbutton(window, text="Headless", variable=headless_var)
headless_checkbox.grid(row=0, column=4, padx=10, pady=50, sticky="W")

status_label = ttk.Label(window, text="Status: Idle")
status_label.configure(foreground="#e74c3c")
status_label.grid(row=1, column=0, columnspan=2, padx=10, pady=30, sticky="W")

# Create the start button with updated color
# start button that call the main function when clicked and disable itself to prevent multiple clicks
start_button = ttk.Button(window, text="Start", command=start_main_thread)
start_button.grid(row=1, column=1, columnspan=2, padx=10, pady=30)


scraping = False  # flag to prevent multiple scraping processes

ff_path = ''
try:
    with open('ff_path.txt', 'r') as file:
        ff_path = file.readline()
        ff_path = ff_path.strip()
        ff_path = ff_path.replace('\n', '')
except Exception as e:
    messagebox.showerror("Error", "Firefox path not found.\n" + str(e))
    exit()


driver = webdriver

# Start the Tkinter event loop
window.mainloop()

